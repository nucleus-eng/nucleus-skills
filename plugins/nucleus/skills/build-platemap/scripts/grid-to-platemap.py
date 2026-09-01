#!/usr/bin/env python3
"""Turn a spatial plate grid into platemap rows, and cross-check any table.

Scientists lay plates out the way the plate looks: plate rows down the left,
plate columns across the top, well contents in the cells. That is a picture
of the plate, not a platemap -- the CDK needs one row per well.

This reads the grid, emits `Well` and `Name` rows, and -- when the same sheet
also holds a written platemap table -- compares the two. Wells present in one
and not the other are the finding worth having: a grid is what was done at
the bench, and a table missing half of it will silently analyse half the run.

    grid-to-platemap.py sheet.xlsx -o wells.tsv

Exit codes: 0 wrote the file, 1 wrote it with cross-check findings, 2 no grid.
"""

from __future__ import annotations

import argparse
import csv
import datetime
import re
import sys
from pathlib import Path

REQUIRED = ["Date", "Experiment", "Well", "Name", "Type", "Rxn Volume (uL)"]
PLATE_ROW = re.compile(r"^[A-Z]{1,2}$")


def is_blank(value) -> bool:
    return value is None or str(value).strip() == ""


def text(value) -> str:
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    return "" if value is None else str(value).strip()


def load_grid(path: Path, sheet: str | None) -> list[list]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            import openpyxl
        except ImportError:
            print("error: reading .xlsx needs openpyxl (pip install openpyxl)", file=sys.stderr)
            raise SystemExit(2)
        book = openpyxl.load_workbook(path, data_only=True)
        worksheet = book[sheet] if sheet else book.worksheets[0]
        return [
            [worksheet.cell(r, c).value for c in range(1, worksheet.max_column + 1)]
            for r in range(1, worksheet.max_row + 1)
        ]
    delimiter = "\t" if path.suffix.lower() in {".tsv", ".tab"} else ","
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return [list(row) for row in csv.reader(handle, delimiter=delimiter)]


def as_int(value):
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def find_grid(rows: list[list]) -> tuple[int, dict[int, int]] | None:
    """Find the plate-column header row and which sheet column holds which plate column."""
    for index, row in enumerate(rows):
        columns = {c: as_int(v) for c, v in enumerate(row) if not is_blank(v)}
        numbers = {c: n for c, n in columns.items() if n is not None and 1 <= n <= 48}
        if len(numbers) >= 3 and len(numbers) == len(columns):
            # The next row must start with a plate-row letter for this to be a grid.
            if index + 1 < len(rows):
                first = text(rows[index + 1][0])
                if PLATE_ROW.match(first):
                    return index, numbers
    return None


def read_grid_wells(rows, header_index, columns) -> list[tuple[str, str]]:
    wells = []
    for row in rows[header_index + 1:]:
        label = text(row[0]) if row else ""
        if not PLATE_ROW.match(label):
            if wells:
                break
            continue
        for sheet_column, plate_column in sorted(columns.items(), key=lambda kv: kv[1]):
            value = text(row[sheet_column]) if sheet_column < len(row) else ""
            if value:
                wells.append((f"{label}{plate_column}", value))
    return wells


def find_table(rows: list[list]) -> tuple[list[str], list[dict[str, str]]] | None:
    for index, row in enumerate(rows):
        labels = [text(c) for c in row]
        if sum(1 for column in REQUIRED if column in labels) >= 3:
            header = labels
            out = []
            for line in rows[index + 1:]:
                if all(is_blank(v) for v in line):
                    break
                out.append({name: text(line[i]) if i < len(line) else ""
                            for i, name in enumerate(header) if name})
            return [h for h in header if h], out
    return None


def normalise(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", name.lower()).strip()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("source", type=Path)
    parser.add_argument("--sheet", help="worksheet name for .xlsx input")
    parser.add_argument("-o", "--output", type=Path, help="write grid wells to this .csv or .tsv")
    args = parser.parse_args()

    if not args.source.is_file():
        print(f"error: no such file: {args.source}", file=sys.stderr)
        return 2

    rows = load_grid(args.source, args.sheet)
    found = find_grid(rows)
    if found is None:
        print("error: no plate grid found (need a row of plate-column numbers "
              "above rows starting with a plate-row letter)", file=sys.stderr)
        return 2
    header_index, columns = found
    grid_wells = read_grid_wells(rows, header_index, columns)
    print(f"grid: {len(grid_wells)} wells across rows "
          f"{', '.join(sorted({w[0][0] for w in grid_wells}))}")

    findings = []

    seen = {}
    for well, contents in grid_wells:
        seen.setdefault(normalise(contents), []).append(well)
    for contents, wells in seen.items():
        if len(wells) > 1 and "replicate" in contents:
            findings.append(
                f"grid wells {', '.join(wells)} carry identical text but name a replicate "
                f"-- one of them is probably a copy-paste of the other"
            )

    table = find_table(rows)
    if table:
        _, table_rows = table
        table_wells = {r.get("Well", "").strip(): r for r in table_rows if r.get("Well", "").strip()}
        grid_only = [w for w, _ in grid_wells if w not in table_wells]
        table_only = [w for w in table_wells if w not in {g for g, _ in grid_wells}]
        print(f"table: {len(table_wells)} wells")
        if grid_only:
            findings.append(
                f"{len(grid_only)} well(s) are in the grid but not in the table "
                f"({', '.join(grid_only[:8])}{', ...' if len(grid_only) > 8 else ''}) -- "
                f"the merge will drop them and analyse only the rest"
            )
        if table_only:
            findings.append(f"{len(table_only)} well(s) are in the table but not the grid: {', '.join(table_only)}")
        for well, contents in grid_wells:
            row = table_wells.get(well)
            if row and normalise(contents) != normalise(row.get("Name", "")):
                findings.append(f"{well}: grid says {contents!r}, table says {row.get('Name')!r}")
    else:
        print("table: none found on this sheet")

    if args.output:
        delimiter = "\t" if args.output.suffix.lower() in {".tsv", ".tab"} else ","
        with args.output.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle, delimiter=delimiter)
            writer.writerow(REQUIRED)
            for well, contents in grid_wells:
                writer.writerow(["", "", well, contents, "", ""])
        print(f"wrote {args.output} -- Date, Experiment, Type and Rxn Volume are blank "
              f"and must be filled in before this is a valid platemap")

    if findings:
        print(f"\n{len(findings)} cross-check finding(s):")
        for finding in findings:
            print(f"  [warn    ] {finding}")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
