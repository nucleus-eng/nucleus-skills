"""Shared definitions for the platemap scripts.

The column and type definitions are owned by the original DevNote,
`nucleus-devnote-archive-1/devnotes/2026-bhasin-platemaps/main.md`. This is a port --
keep it in sync with that page. It lives here, once, because three scripts
need it, and three copies of a definition is how two copies come to disagree.

Where the published tutorial at
docs.nucleus.engineering/guides/platemap-tutorial/ disagrees with the
DevNote, the DevNote wins. It differs in two places:

- the tutorial makes `Rxn Volume (uL)` a sixth required column; the DevNote
  lists reaction volume among the optional ones. It is strongly recommended
  here, and its absence is a warning rather than an error.
- the tutorial omits `Blank` from the type vocabulary. The DevNote lists it,
  and the CDK's `blank_data()` uses it as its default `blank_type`, so the
  tutorial is the outlier.
"""

from __future__ import annotations

import csv
from pathlib import Path

RXN_VOLUME = "Rxn Volume (uL)"

# Required: absence is an error. Five columns, per the DevNote.
REQUIRED = ["Date", "Experiment", "Well", "Name", "Type"]
# Recommended: absence is a warning. Analysis runs; the record is poorer.
RECOMMENDED = [RXN_VOLUME]
# What a platemap this repo writes should carry.
PLATEMAP_COLUMNS = REQUIRED + RECOMMENDED

TYPES = {"Sample", "Standard", "Blank",
         "Control", "Positive Control", "Negative Control"}
# DEFAULT_ANALYSIS_COLUMNS in the CDK's platereader.py -- kinetics runs on these only.
ANALYSED = {"Sample", "Control", "Positive Control"}

# Last row letter and last column, per plate format. 384 is the default here.
PLATES = {96: ("H", 12), 384: ("P", 24), 1536: ("AF", 48)}


def plate_rows(last_row: str) -> list[str]:
    """Row labels up to `last_row`: A..Z, then AA..AF for 1536."""
    single = [chr(c) for c in range(ord("A"), ord("Z") + 1)]
    labels = single + [f"A{c}" for c in single]
    return labels[: labels.index(last_row) + 1]


def is_blank(value) -> bool:
    return value is None or str(value).strip() == ""


def load_grid(path: Path, sheet: str | None = None) -> list[list]:
    """Read a sheet as a raw grid of cells. Handles .xlsx, .csv and .tsv."""
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            import openpyxl
        except ImportError as exc:
            raise SystemExit("error: reading .xlsx needs openpyxl (pip install openpyxl)") from exc
        book = openpyxl.load_workbook(path, data_only=True)
        worksheet = book[sheet] if sheet else book.worksheets[0]
        return [
            [worksheet.cell(r, c).value for c in range(1, worksheet.max_column + 1)]
            for r in range(1, worksheet.max_row + 1)
        ]
    delimiter = "\t" if path.suffix.lower() in {".tsv", ".tab"} else ","
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return [list(row) for row in csv.reader(handle, delimiter=delimiter)]
